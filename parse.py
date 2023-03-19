""" This module parses the data from the excel files.

It have two functions parsing_worksheet1 and parsing_worksheet2 that 
performs parsing of data on different worksheets tabs named as Summary 
Rolling Row and VOC Rolling MoM
"""
from datetime import datetime
import shutil
import os
import openpyxl
from flask import render_template


def parsing_worksheet1(file,file_month,file_year):
    """Parses data from worksheet Summary Rolling MoM.

       It retrieves validated file name,month and year and 
       parses the data from the worksheet tab Summary Rolling MoM.

       Args:
         file: String containing the validated file-name.
         file_month:Integer value of the month for which data has to be parsed.
         file_year:Integer value of  year for which data has to be parsed.

       Returns:
         Calls_Offered :Interger
         Abandon_after_30s :Float
         FCR :Float
         DSAT :Float
         CSAT :Float
       Raises:
        Error When Valid file-name has missing worksheets tabs.
    """
    wb_obj=openpyxl.load_workbook(f"/vagrant/project_flask/ClientFolder/{file}")
    if wb_obj['Summary Rolling MoM'] is None:
        os.remove(f"/vagrant/project_flask/ArchiveFolder/{file}")
        raise Exception('Error')
    sheet_obj = wb_obj['Summary Rolling MoM']
    max_row = sheet_obj.max_row
    list1=[]
    for i in range(1, max_row+ 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        if isinstance(cell_obj.value,datetime):
            if cell_obj.value.year==file_year and cell_obj.value.month==file_month:
                for j in range(2,7):
                    cell_obj1 = sheet_obj.cell(row = i, column = j)
                    list1.append(cell_obj1.value)

    calls_offered=int(list1[0])
    abandon_after_30s=float(list1[1])*100
    f_c_r=float(list1[2])*100
    d_s_a_t=float(list1[3])*100
    c_s_a_t=float(list1[4])*100  
    return calls_offered,abandon_after_30s,f_c_r,d_s_a_t,c_s_a_t

def parsing_worksheet2(file,file_month,file_year):
    """Parses data from worksheet VOC Rolling MoM.

       It retrieves validated file name,month and year and 
       parses the data from the worksheet tab VOC Rolling MoM and 
       further determines whether Promoters,Passives,Dectractors
       are good or bad

       Args:
         file: String containing the validated file-name.
         file_month:Integer value of the month for which data has to be parsed.
         file_year:Integer value of  year for which data has to be parsed.

       Returns:
         Promoters :String
         Passives :String
         Dectractors:String
       Raises:
        Error When Valid file-name has missing worksheets tabs.
    """
    wb_obj=openpyxl.load_workbook(f"/vagrant/project_flask/ClientFolder/{file}")
    if wb_obj['VOC Rolling MoM'] is None:
        os.remove(f"/vagrant/project_flask/ArchiveFolder/{file}")
        raise Exception('Error')
    sheet_obj1 = wb_obj['VOC Rolling MoM']
    cell_obj1 = sheet_obj1.cell(row = 1, column = 2)
    max_col1=sheet_obj1.max_column
    list2=[]
    for i in range(1, max_col1+ 1):
        cell_obj1 = sheet_obj1.cell(row = 1, column = i)
        if  isinstance(cell_obj1.value,datetime):
            if cell_obj1.value.year==file_year and cell_obj1.value.month==file_month:
                for j in range(4,9,2):
                    cell_obj1 = sheet_obj1.cell(row = j, column = i)
                    list2.append(cell_obj1.value)
    pro_moters=int(list2[0])              
    pas_sives=int(list2[1]) 
    dec_tractors=int(list2[2])  
    if pro_moters>200:
        pro_moters="good"
    else:
         pro_moters="bad"         
    if pas_sives>100:
        pas_sives="good"
    else:
        pas_sives="bad"
    if dec_tractors>100:
        dec_tractors="good"
    else:
        dec_tractors="bad"
    return pro_moters,pas_sives,dec_tractors

